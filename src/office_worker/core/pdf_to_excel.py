"""Conversión y extracción de tablas de PDF a Excel (.xlsx) con pdfplumber y openpyxl."""
from __future__ import annotations
import os
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .security import safe_out


def pdf_to_excel(
    input_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    max_pages: int | None = None,
) -> dict[str, Any]:
    """Extrae tablas estructuradas desde un archivo PDF y las guarda en un archivo Excel (.xlsx).

    - input_path: ruta al archivo PDF origen.
    - output_path: ruta destino del archivo .xlsx.
    - sheet_name: nombre de la hoja en el libro Excel.
    - max_pages: límite opcional de páginas a analizar.

    Devuelve dict con fidelity honesta ("clean"|"lossy"), conteo de tablas y warnings.
    """
    import pdfplumber

    input_path = os.path.abspath(os.path.expanduser(str(input_path)))
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"PDF de entrada no encontrado: {input_path}")

    out_path = safe_out(output_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]

    header_fill = PatternFill("solid", fgColor="FF003366")
    header_font = Font(bold=True, color="FFFFFFFF")
    alt_fill = PatternFill("solid", fgColor="FFF5F7FA")
    thin = Side(style="thin", color="FFC8D4E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    tables_found: list[list[list[Any]]] = []
    pages_processed = 0

    with pdfplumber.open(input_path) as pdf:
        limit = len(pdf.pages) if not max_pages else min(len(pdf.pages), int(max_pages))
        pages_processed = limit
        for i in range(limit):
            pg = pdf.pages[i]
            extracted = pg.extract_tables() or []
            for t in extracted:
                if t and any(row and any(c is not None and str(c).strip() for c in row) for row in t):
                    tables_found.append(t)

    current_row = 1
    warnings: list[str] = []

    if not tables_found:
        ws.cell(1, 1, "No structured tables detected in PDF")
        ws.cell(1, 1).font = Font(italic=True, color="FF718096")
        fidelity = "lossy"
        warnings.append("No structured tables detected in PDF; empty sheet created.")
    else:
        fidelity = "clean"
        for t_idx, table in enumerate(tables_found, start=1):
            if t_idx > 1:
                # Línea separadora entre tablas
                current_row += 2

            for r_idx, row in enumerate(table):
                is_header = (r_idx == 0)
                for c_idx, val in enumerate(row, start=1):
                    clean_val = str(val).strip() if val is not None else ""
                    cell = ws.cell(current_row, c_idx, clean_val)
                    cell.border = border
                    if is_header:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="top")
                        if r_idx % 2 == 0:
                            cell.fill = alt_fill
                current_row += 1

        # Autoajuste de anchos de columna (máx 45)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    wb.save(out_path)
    return {
        "status": "ok",
        "path": os.path.abspath(out_path),
        "bytes": os.path.getsize(out_path),
        "n_tables": len(tables_found),
        "pages_processed": pages_processed,
        "fidelity": fidelity,
        "warnings": warnings,
    }
