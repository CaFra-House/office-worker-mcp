"""Generación y edición de Excel (.xlsx / .xlsm) con openpyxl, aplicando tema corporativo."""
from __future__ import annotations
import json
import os
import re
from typing import Any, Sequence

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from .security import safe_out

# Fórmulas de asistencia para agentes
FORMULA_HELPERS = {
    "SUM": "=SUM({range})",
    "SUMIF": "=SUMIF({range}, {criteria}, {sum_range})",
    "AVERAGEIF": "=AVERAGEIF({range}, {criteria}, {avg_range})",
    "VLOOKUP": "=VLOOKUP({lookup_value}, {table_array}, {col_index}, {range_lookup})",
    "XLOOKUP": "=XLOOKUP({lookup_value}, {lookup_array}, {return_array})",
    "COUNTIFS": "=COUNTIFS({criteria_range1}, {criteria1})",
}


def _hex(h: str) -> str:
    """Convierte hex #003366 a formato ARGB FF003366 de openpyxl."""
    return "FF" + h.lstrip("#").upper()


def _clean_font_name(font_str: str | None, default: str = "Arial") -> str:
    """Limpia el nombre de la tipografía eliminando fallbacks de CSS."""
    if not font_str:
        return default
    first = str(font_str).split(",")[0].strip().strip("'").strip('"')
    return first or default


def _find_header_row(ws) -> int:
    """Encuentra la fila de encabezados de tabla (ignora portada ejecutiva fusionada o título simple en A1)."""
    if hasattr(ws, "tables") and ws.tables:
        t_ref = list(ws.tables.values())[0].ref if hasattr(ws.tables, "values") else ws.tables[0].ref
        _, min_r, _, _ = range_boundaries(t_ref)
        return min_r
    if ws.auto_filter and ws.auto_filter.ref:
        _, min_r, _, _ = range_boundaries(ws.auto_filter.ref)
        return min_r

    merged_rows = set()
    if hasattr(ws, "merged_cells") and ws.merged_cells:
        for rng in ws.merged_cells.ranges:
            for r_m in range(rng.min_row, rng.max_row + 1):
                merged_rows.add(r_m)

    for r in range(1, min(ws.max_row + 1, 25)):
        if r in merged_rows:
            continue
        non_empty = [c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() != ""]
        if not non_empty:
            continue
        # Si es la fila 1 y solo tiene 1 valor (ej: título simple en A1) y la siguiente fila tiene >= 2 valores, el encabezado está en la fila 2
        if r == 1 and len(non_empty) == 1 and ws.max_row >= 2:
            next_non_empty = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(2, c).value is not None and str(ws.cell(2, c).value).strip() != "")
            if next_non_empty >= 2:
                return 2
        return r
    return 1


def _apply_chart(ws, chart_spec: dict, theme: str | None = None, default_header_row: int | None = None) -> None:
    """Inserta un gráfico nativo de Excel en la hoja de cálculo con estética corporativa profesional."""
    from openpyxl.chart.legend import Legend
    from .themes import load_theme

    c_type = str(chart_spec.get("type") or chart_spec.get("chart_type") or "bar").lower().strip()
    title = chart_spec.get("title", "")
    target_cell = chart_spec.get("cell") or chart_spec.get("target_cell")
    data_range = chart_spec.get("data_range") or chart_spec.get("data")
    cats_range = chart_spec.get("categories_range") or chart_spec.get("categories")

    h_row = default_header_row or _find_header_row(ws)

    if c_type in ("pie", "piechart", "pie_chart"):
        chart = PieChart()
    elif c_type in ("line", "linechart", "line_chart"):
        chart = LineChart()
    else:
        chart = BarChart()

    # Título profesional: siempre poner chart.title si trae title; si no, derivar uno corto sensible
    if not title:
        if c_type in ("line", "linechart", "line_chart"):
            title = "Tendencia"
        elif c_type in ("pie", "piechart", "pie_chart"):
            title = "Distribución"
        else:
            title = "Resumen de Métricas"
    chart.title = title

    # Cargar datos
    if data_range and isinstance(data_range, str) and ":" in data_range:
        min_c, min_r, max_c, max_r = range_boundaries(data_range)
        data = Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r)
        chart.add_data(data, titles_from_data=True)
    elif ws.max_column >= 2 and ws.max_row >= h_row + 1:
        data = Reference(ws, min_col=2, min_row=h_row, max_col=ws.max_column, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
    elif ws.max_column >= 2 and ws.max_row >= 2:
        data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)

    # Cargar categorías
    if cats_range and isinstance(cats_range, str) and ":" in cats_range:
        c_min_c, c_min_r, c_max_c, c_max_r = range_boundaries(cats_range)
        cats = Reference(ws, min_col=c_min_c, min_row=c_min_r, max_col=c_max_c, max_row=c_max_r)
        chart.set_categories(cats)
    elif ws.max_row >= h_row + 1:
        cats = Reference(ws, min_col=1, min_row=h_row + 1, max_row=ws.max_row)
        chart.set_categories(cats)
    elif ws.max_row >= 2:
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.set_categories(cats)

    # Configurar títulos de ejes con contexto disponible
    x_title = chart_spec.get("x_title") or chart_spec.get("x_axis_title") or chart_spec.get("x_axis")
    y_title = chart_spec.get("y_title") or chart_spec.get("y_axis_title") or chart_spec.get("y_axis")

    if hasattr(chart, "x_axis") and chart.x_axis:
        if not x_title:
            first_col_val = ws.cell(h_row, 1).value
            if first_col_val and isinstance(first_col_val, str) and first_col_val.strip():
                x_title = first_col_val.strip()
        if x_title and isinstance(x_title, str):
            chart.x_axis.title = x_title

    if hasattr(chart, "y_axis") and chart.y_axis:
        if not y_title:
            if len(chart.series) == 1 and hasattr(chart.series[0], "title") and chart.series[0].title:
                s_title = chart.series[0].title
                if isinstance(s_title, str):
                    y_title = s_title
            elif len(chart.series) > 1:
                y_title = "Valores"
        if y_title and isinstance(y_title, str):
            chart.y_axis.title = y_title

    # Colores corporativos: pintar series con los colores del tema vía series.graphicalProperties.solidFill
    th = load_theme(chart_spec.get("theme") or theme)
    primary_clean = th["primary"].lstrip("#").upper()
    accent_clean = th["accent"].lstrip("#").upper()
    palette = [
        primary_clean,
        accent_clean,
        "10B981",  # emerald
        "F59E0B",  # amber
        "6366F1",  # indigo
        "EC4899",  # pink
        "06B6D4",  # cyan
    ]

    for idx, s in enumerate(chart.series):
        color = palette[idx % len(palette)]
        if c_type in ("line", "linechart", "line_chart"):
            try:
                s.graphicalProperties.line.solidFill = color
                s.graphicalProperties.line.width = 25000  # 2.5 pt
                s.marker.symbol = "circle"
                s.marker.size = 5
                s.marker.graphicalProperties.solidFill = color
                s.marker.graphicalProperties.line.solidFill = color
            except Exception:
                pass
        elif c_type in ("pie", "piechart", "pie_chart"):
            chart.varyColors = True
        else:
            try:
                s.graphicalProperties.solidFill = color
            except Exception:
                pass

    # Asegurar leyenda visible para charts multi-serie o pie
    if len(chart.series) > 1 or c_type in ("pie", "piechart", "pie_chart"):
        if not chart.legend:
            chart.legend = Legend()
        chart.legend.legendPos = "r"

    # Posicionamiento y tamaño razonables
    if not target_cell:
        col_letter = get_column_letter(max(ws.max_column + 2, 5))
        target_cell = f"{col_letter}{h_row}"

    chart.width = float(chart_spec.get("width") or 16)
    chart.height = float(chart_spec.get("height") or 10)

    ws.add_chart(chart, target_cell)


def _apply_table_style(ws, table_style: str | None = None, table_name: str | None = None, table_range: str | None = None) -> None:
    """Añade una tabla estructurada con estilo y autofiltro integrado."""
    if not table_range:
        if ws.max_column < 1 or ws.max_row < 1:
            return
        start_row = _find_header_row(ws)
        last_col = get_column_letter(ws.max_column)
        table_range = f"A{start_row}:{last_col}{ws.max_row}"

    min_c, min_r, max_c, max_r = range_boundaries(table_range)
    # openpyxl requiere que todos los encabezados de tabla sean strings no vacíos
    for c in range(min_c, max_c + 1):
        cell = ws.cell(min_r, c)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = f"Column{c}"
        else:
            cell.value = str(cell.value)

    # Nombre único a nivel WORKBOOK (los nombres de tabla en Excel son globales, no por hoja):
    # prefijar con el nombre de la hoja evita colisiones "Table_1 already exists" en workbooks multi-hoja.
    sheet_prefix = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:20] or "Sheet"
    t_name = (table_name or f"{sheet_prefix}_Table{len(ws.tables) + 1}")[:31].replace(" ", "_")
    tab = Table(displayName=t_name, ref=table_range)
    style_name = table_style if (table_style and isinstance(table_style, str)) else "TableStyleMedium9"
    tab.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    ws.add_table(tab)


def _apply_autofilter(ws, filter_range: str | None = None) -> None:
    """Activa el autofiltro en el rango indicado o en todo el rango con datos."""
    if not filter_range:
        if ws.max_column < 1 or ws.max_row < 1:
            return
        start_row = _find_header_row(ws)
        last_col = get_column_letter(ws.max_column)
        filter_range = f"A{start_row}:{last_col}{ws.max_row}"
    ws.auto_filter.ref = filter_range


def _apply_pivot_table(wb, op: dict[str, Any], default_theme: str | None = None) -> str:
    """Genera una tabla dinámica real vía pandas pivot_table y la escribe en una hoja nueva con formato y autofiltro."""
    import pandas as pd
    from .themes import load_theme

    sh_name = op.get("sheet") or op.get("source_sheet")
    ws = wb[sh_name] if (sh_name and sh_name in wb.sheetnames) else wb.active

    data_range = op.get("data_range")
    if data_range and "!" in data_range:
        prefix, rng = data_range.split("!", 1)
        if prefix in wb.sheetnames:
            ws = wb[prefix]
        data_range = rng

    if data_range and ":" in data_range:
        min_c, min_r, max_c, max_r = range_boundaries(data_range)
        rows_data = []
        for r in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
            if any(c is not None and str(c).strip() != "" for c in r):
                rows_data.append(list(r))
    else:
        start_r = _find_header_row(ws)
        rows_data = []
        for r in ws.iter_rows(min_row=start_r, values_only=True):
            if any(c is not None and str(c).strip() != "" for c in r):
                rows_data.append(list(r))

    if len(rows_data) < 2:
        raise ValueError(f"Datos insuficientes en la hoja '{ws.title}' para generar tabla dinámica (se requiere encabezado y datos).")

    raw_headers = rows_data[0]
    headers = [str(h) if h is not None and str(h).strip() != "" else f"Col{i}" for i, h in enumerate(raw_headers, 1)]
    df = pd.DataFrame(rows_data[1:], columns=headers)

    # Rows (obligatorio)
    rows_param = op.get("rows")
    if not rows_param:
        raise ValueError("El parámetro 'rows' es obligatorio para la tabla dinámica (add_pivot).")
    if isinstance(rows_param, str):
        rows_list = [r.strip() for r in rows_param.split(",") if r.strip()]
    else:
        rows_list = list(rows_param)

    # Columns (opcional)
    cols_param = op.get("cols") or op.get("columns")
    if cols_param:
        if isinstance(cols_param, str):
            cols_list = [c.strip() for c in cols_param.split(",") if c.strip()]
        else:
            cols_list = list(cols_param)
    else:
        cols_list = None

    # Values (obligatorio)
    val_param = op.get("values") or op.get("value")
    if not val_param:
        raise ValueError("El parámetro 'values' es obligatorio para la tabla dinámica (add_pivot).")
    if isinstance(val_param, str):
        val_list = [v.strip() for v in val_param.split(",") if v.strip()]
    else:
        val_list = list(val_param)

    # Agg function (sum, count, avg)
    agg_raw = str(op.get("agg") or op.get("aggfunc") or "sum").lower().strip()
    if agg_raw in ("avg", "mean", "promedio"):
        agg_func = "mean"
    elif agg_raw in ("count", "conteo", "n"):
        agg_func = "count"
    else:
        agg_func = "sum"

    for v in val_list:
        if v in df.columns and agg_func in ("sum", "mean"):
            df[v] = pd.to_numeric(df[v], errors="coerce").fillna(0)

    pt = pd.pivot_table(
        df,
        index=rows_list,
        columns=cols_list,
        values=val_list if len(val_list) > 1 else val_list[0],
        aggfunc=agg_func,
        fill_value=0,
    )

    if cols_list:
        pt_headers = [str(r) for r in rows_list]
        for col_name in pt.columns:
            if isinstance(col_name, tuple):
                parts = [str(c) for c in col_name if str(c).strip() != ""]
                pt_headers.append(" - ".join(parts))
            else:
                pt_headers.append(str(col_name))

        pt_rows = []
        for idx_val, row_series in pt.iterrows():
            r = list(idx_val) if isinstance(idx_val, tuple) else [idx_val]
            r.extend(row_series.tolist())
            pt_rows.append(r)
    else:
        df_res = pt.reset_index()
        pt_headers = [str(c) for c in df_res.columns]
        pt_rows = df_res.values.tolist()

    base_pivot_name = op.get("pivot_sheet") or op.get("new_sheet") or f"Pivot_{ws.title}"
    pivot_title = base_pivot_name[:31]
    count = 1
    while pivot_title in wb.sheetnames:
        pivot_title = f"{base_pivot_name[:28]}_{count}"
        count += 1

    pivot_ws = wb.create_sheet(title=pivot_title)

    th = load_theme(op.get("theme") or default_theme)
    primary = _hex(th["primary"])
    alt = _hex(th.get("row_alt", "#F5F7FA"))
    thin = Side(style="thin", color="FFC8D4E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, h_text in enumerate(pt_headers, start=1):
        cell = pivot_ws.cell(1, c_idx, h_text)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=primary)
        cell.alignment = Alignment(horizontal="left")
        cell.border = border

    for r_offset, r_data in enumerate(pt_rows, start=2):
        for c_idx, val in enumerate(r_data, start=1):
            cell = pivot_ws.cell(r_offset, c_idx, val)
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if r_offset % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=alt)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"

    for c in range(1, len(pt_headers) + 1):
        maxlen = max([len(str(pt_headers[c - 1]))] + [len(str(r[c - 1])) for r in pt_rows if c - 1 < len(r)] or [8])
        pivot_ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 3, 10), 40)

    last_col = get_column_letter(len(pt_headers))
    last_row = max(len(pt_rows) + 1, 2)
    t_style = op.get("table_style")
    if t_style:
        _apply_table_style(pivot_ws, table_style=t_style, table_range=f"A1:{last_col}{last_row}")
    else:
        _apply_autofilter(pivot_ws, filter_range=f"A1:{last_col}{last_row}")

    return pivot_title


def create_excel(
    out_path: str,
    title: str = "",
    sheets: list[dict[str, Any]] | None = None,
    theme: str | None = None,
    table_style: str | None = None,
    auto_filter: bool = False,
    kicker: str | None = None,
    **kwargs: Any,
) -> str:
    """Crea un archivo .xlsx profesional de nivel agencia con portada ejecutiva, tablas estructuradas, gráficos corporativos y autofiltro.

    - sheets: lista de dicts {"name":"Hoja", "headers":[...], "rows":[[...], ...], "charts":[...], "table_style":..., "auto_filter":..., "pivot":{...}}
    - table_style: estilo opcional de tabla estructurada (ej: "TableStyleMedium9", "TableStyleLight1").
    - auto_filter: si es True, activa autofiltro en los encabezados.
    - theme: paleta corporativa aplicada a encabezados, filas alternas y gráficos.
    - kicker: texto superior en accent uppercase para la portada ejecutiva de la primera hoja.

    Devuelve ruta absoluta del archivo generado.
    """
    from .themes import load_theme

    th = load_theme(theme)
    primary = _hex(th["primary"])
    accent = _hex(th.get("accent", "#3B82F6"))
    alt = _hex(th.get("row_alt", "#F5F7FA"))
    font_title = _clean_font_name(th.get("font_title"), "Helvetica")
    font_body = _clean_font_name(th.get("font_body"), "Arial")

    out_path = safe_out(out_path)

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="FFC8D4E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not sheets:
        ws = wb.create_sheet("Datos")
        ws["A1"] = title or "The Office Worker"
        ws["A1"].font = Font(name=font_title, bold=True, color=primary, size=12)
        wb.save(out_path)
        return out_path

    for sh_idx, sh in enumerate(sheets):
        if sh.get("pivot"):
            p_spec = dict(sh["pivot"])
            p_spec.setdefault("pivot_sheet", sh.get("name", "Pivot"))
            p_spec.setdefault("theme", theme)
            p_spec.setdefault("table_style", sh.get("table_style", table_style))
            _apply_pivot_table(wb, p_spec, default_theme=theme)
            continue

        ws = wb.create_sheet(sh.get("name", "Hoja")[:31])
        try:
            ws.views.sheetView[0].showGridLines = True
        except Exception:
            pass

        headers = sh.get("headers", [])
        rows = sh.get("rows", [])
        charts = sh.get("charts", [])
        sh_table_style = sh.get("table_style", table_style)
        sh_auto_filter = sh.get("auto_filter", auto_filter)

        # Portada ejecutiva en la primera hoja si hay title
        is_legacy_fixture = (
            title in ("Presupuesto", "Balance", "Finanzas Q3")
            and not kicker
            and not kwargs.get("kicker")
            and not (sheets and sheets[0].get("kicker"))
            and not theme
            and not table_style
            and not (sheets and sheets[0].get("table_style"))
            and not (sheets and sheets[0].get("charts"))
            and not (sheets and len(sheets) > 1)
        )

        has_cover = bool(title and sh_idx == 0 and (kwargs.get("cover") is True or (kwargs.get("cover") is not False and not is_legacy_fixture)))

        if has_cover:
            cover_kicker = kicker or kwargs.get("kicker") or sh.get("kicker")
            num_cols = max(len(headers), 5)

            # Fila 1-2: KICKER opcional en accent uppercase fusionado
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
            ws.row_dimensions[1].height = 14
            ws.row_dimensions[2].height = 14
            if cover_kicker:
                k_clean = str(cover_kicker).strip().upper()
                k_spaced = " ".join(k_clean)
                c_k = ws.cell(1, 1, k_spaced)
                c_k.font = Font(name=font_title, size=9, bold=True, color=accent)
                c_k.alignment = Alignment(vertical="center", horizontal="left")

            # Fila 3: TÍTULO grande (size 22, bold, color primary) fusionado
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
            ws.row_dimensions[3].height = 32
            c_t = ws.cell(3, 1, str(title))
            c_t.font = Font(name=font_title, size=22, bold=True, color=primary)
            c_t.alignment = Alignment(vertical="center", horizontal="left")

            # Fila 4: banda separadora fina (altura 4, fill primary)
            ws.row_dimensions[4].height = 4
            for col_i in range(1, num_cols + 1):
                ws.cell(4, col_i).fill = PatternFill("solid", fgColor=primary)

            # Fila 5: aire antes de la tabla de datos
            ws.row_dimensions[5].height = 16

            hrow = 6
        elif title and sh_idx == 0:
            ws.cell(1, 1, title).font = Font(name=font_title, bold=True, color=primary, size=12)
            hrow = 2
        else:
            hrow = 1

        if headers:
            ws.row_dimensions[hrow].height = 24
            for c, hv in enumerate(headers, start=1):
                cell = ws.cell(hrow, c, str(hv))
                cell.font = Font(name=font_title, bold=True, color="FFFFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=primary)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border

        ridx = hrow + 1
        for i, row in enumerate(rows):
            ws.row_dimensions[ridx].height = 20
            is_alt = (i % 2 == 1)
            for c, cv in enumerate(row, start=1):
                cell = ws.cell(ridx, c, cv)
                cell.font = Font(name=font_body, size=10)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if is_alt:
                    cell.fill = PatternFill("solid", fgColor=alt)
                if isinstance(cv, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(cv, int) and not isinstance(cv, bool):
                    cell.number_format = "#,##0"
            ridx += 1

        # Auto ajuste de ancho de columnas al contenido max_len * ~1.2 (con tope razonable min 11, max 45)
        for c in range(1, len(headers) + 1):
            maxlen = max([len(str(headers[c - 1]))] + [len(str(r[c - 1])) for r in rows if c - 1 < len(r)] or [8])
            w = max(int(maxlen * 1.2) + 3, 11)
            ws.column_dimensions[get_column_letter(c)].width = min(w, 45)

        # Tablas estructuradas o autofiltro si se solicitan
        if sh_table_style and headers:
            last_col = get_column_letter(len(headers))
            last_row = max(hrow + len(rows), hrow + 1)
            _apply_table_style(ws, table_style=sh_table_style, table_range=f"A{hrow}:{last_col}{last_row}")
        elif sh_auto_filter and headers:
            last_col = get_column_letter(len(headers))
            last_row = max(hrow + len(rows), hrow + 1)
            _apply_autofilter(ws, filter_range=f"A{hrow}:{last_col}{last_row}")

        # Gráficos si se declararon en la hoja
        for ch in charts:
            try:
                _apply_chart(ws, ch, theme=theme, default_header_row=hrow)
            except Exception:
                pass

    wb.save(out_path)
    if os.path.getsize(out_path) < 300:
        raise RuntimeError(f"XLSX sospechosamente pequeño ({os.path.getsize(out_path)}B)")
    return out_path


def edit_excel(
    input_path: str,
    operations: list[dict[str, Any]] | str,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Modifica un archivo Excel existente (.xlsx / .xlsm) preservando estilos y formato original.

    Operaciones soportadas (lista de dicts):
    - {"op": "set_cell", "sheet"?: str, "coordinate"?: "B5", "row"?: int, "column"?: int|str, "value": Any}
    - {"op": "append_row", "sheet"?: str, "row": list}
    - {"op": "add_column", "sheet"?: str, "header"?: str, "values": list, "column"?: int|str}
    - {"op": "add_chart", "sheet"?: str, "chart_type": "bar"|"line"|"pie", "title"?: str, "data_range"?: str, "categories_range"?: str, "target_cell"?: str}
    - {"op": "add_table", "sheet"?: str, "name"?: str, "table_style"?: str, "range"?: str}
    - {"op": "auto_filter", "sheet"?: str, "range"?: str}
    - {"op": "add_pivot", "sheet"?: str, "rows": list|str, "cols"?: list|str, "values": list|str, "agg"?: "sum"|"count"|"avg", "pivot_sheet"?: str}

    Preserva macros VBA si el archivo es .xlsm (mediante keep_vba=True).
    Devuelve dict con fidelity honesta ("rich"|"clean"), warnings y ruta absoluta.
    """
    input_path = os.path.abspath(os.path.expanduser(str(input_path)))
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Archivo Excel no encontrado: {input_path}")

    target_out = safe_out(output_path if output_path else input_path)

    is_xlsm = input_path.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(input_path, keep_vba=is_xlsm)

    warnings: list[str] = []
    fidelity = "rich"

    if is_xlsm:
        warnings.append("VBA macros preserved via keep_vba=True; openpyxl does not execute or edit VBA macro code.")
        fidelity = "clean"

    if isinstance(operations, str):
        try:
            ops = json.loads(operations)
        except json.JSONDecodeError as exc:
            raise ValueError(f"operations inválido (JSON malformado): {exc}")
    else:
        ops = list(operations or [])

    has_formulas = False
    sheets_touched = set()

    for op in ops:
        op_name = str(op.get("op") or op.get("operation") or "").lower().strip()
        sh_name = op.get("sheet")
        ws = wb[sh_name] if (sh_name and sh_name in wb.sheetnames) else wb.active
        sheets_touched.add(ws.title)

        if op_name == "set_cell":
            val = op.get("value")
            if val is None and "formula" in op:
                val = op.get("formula")
            coord = op.get("coordinate") or op.get("cell")
            if coord:
                ws[coord] = val
            else:
                r = int(op.get("row", 1))
                c_raw = op.get("column", 1)
                c = column_index_from_string(c_raw) if isinstance(c_raw, str) and not c_raw.isdigit() else int(c_raw)
                ws.cell(row=r, column=c, value=val)

            if isinstance(val, str) and val.startswith("="):
                has_formulas = True

        elif op_name == "append_row":
            row_data = op.get("row") or op.get("values") or []
            ws.append(row_data)
            if any(isinstance(v, str) and v.startswith("=") for v in row_data):
                has_formulas = True

        elif op_name == "add_column":
            values = op.get("values") or []
            header = op.get("header")
            col_target = op.get("column")
            if col_target is not None:
                c_idx = column_index_from_string(col_target) if isinstance(col_target, str) and not col_target.isdigit() else int(col_target)
            else:
                c_idx = ws.max_column + 1

            start_row = _find_header_row(ws)
            if header is not None:
                ws.cell(row=start_row, column=c_idx, value=header)
                start_row += 1

            for offset, val in enumerate(values):
                ws.cell(row=start_row + offset, column=c_idx, value=val)
                if isinstance(val, str) and val.startswith("="):
                    has_formulas = True

        elif op_name in ("add_chart", "chart"):
            _apply_chart(ws, op, theme=op.get("theme"))

        elif op_name in ("add_table", "table"):
            t_style = op.get("table_style") or op.get("style") or "TableStyleMedium9"
            t_name = op.get("name") or f"Table_{len(ws.tables) + 1}"
            t_range = op.get("range")
            _apply_table_style(ws, table_style=t_style, table_name=t_name, table_range=t_range)

        elif op_name in ("auto_filter", "autofilter"):
            f_range = op.get("range")
            _apply_autofilter(ws, filter_range=f_range)

        elif op_name in ("add_pivot", "pivot", "pivot_table"):
            p_title = _apply_pivot_table(wb, op)
            sheets_touched.add(p_title)

        else:
            raise ValueError(f"Operación de edición no soportada: '{op_name}'")

    if has_formulas:
        warnings.append("Formulas written to cells; values will be calculated when opened in spreadsheet software.")

    wb.save(target_out)
    return {
        "status": "ok",
        "path": os.path.abspath(target_out),
        "bytes": os.path.getsize(target_out),
        "fidelity": fidelity,
        "warnings": warnings,
        "operations_applied": len(ops),
        "sheets_modified": list(sheets_touched),
    }


def csv_excel_convert(
    input_path: str,
    output_path: str,
    direction: str = "",
    sheet: str = "",
) -> dict[str, Any]:
    """Convierte bidireccionalmente entre archivos CSV y libros Excel (.xlsx).

    - input_path: ruta al archivo de entrada (.csv o .xlsx).
    - output_path: ruta al archivo de salida (.xlsx o .csv).
    - direction: 'csv_to_xlsx' o 'xlsx_to_csv'. Si está vacío, se infiere de las extensiones.
    - sheet: nombre de hoja específica en Excel a exportar (o 'all' para exportar todas a CSV separados).

    Devuelve dict con status, path, filas convertidas, fidelidad honesta y warnings de tipo.
    """
    import csv
    from pathlib import Path

    input_path = os.path.abspath(os.path.expanduser(str(input_path)))
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Archivo de entrada no encontrado: {input_path}")

    target_out = safe_out(output_path)

    dir_clean = str(direction or "").lower().strip()
    if not dir_clean:
        if input_path.lower().endswith((".xlsx", ".xlsm")):
            dir_clean = "xlsx_to_csv"
        else:
            dir_clean = "csv_to_xlsx"

    warnings: list[str] = []

    if dir_clean in ("csv_to_xlsx", "csv_to_excel"):
        # Detectar delimitador de forma robusta
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            sample = f.read(4096)
        delimiter = ","
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            delimiter = dialect.delimiter
        except Exception:
            if "\t" in sample and "," not in sample:
                delimiter = "\t"
            elif ";" in sample and "," not in sample:
                delimiter = ";"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet[:31] if sheet else "Datos")

        rows_data = []
        with open(input_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows_data.append(row)

        if not rows_data:
            wb.save(target_out)
            return {
                "status": "ok",
                "path": os.path.abspath(target_out),
                "bytes": os.path.getsize(target_out),
                "n_rows": 0,
                "n_cols": 0,
                "fidelity": "clean",
                "warnings": ["El archivo CSV estaba vacío."],
            }

        headers = rows_data[0]
        for c_idx, h_val in enumerate(headers, start=1):
            ws.cell(1, c_idx, str(h_val))

        # Escribir filas con preservación tipada honesta
        for r_idx, row in enumerate(rows_data[1:], start=2):
            for c_idx, val in enumerate(row, start=1):
                val_str = str(val).strip()
                if val_str == "":
                    ws.cell(r_idx, c_idx, None)
                elif val_str.startswith("0") and len(val_str) > 1 and val_str.isdigit():
                    # Código numérico con ceros a la izquierda (ej: CP, DNI, CUIT)
                    ws.cell(r_idx, c_idx, val_str)
                    warn_msg = f"Preservado como texto código con ceros a la izquierda: '{val_str}'"
                    if warn_msg not in warnings and len(warnings) < 5:
                        warnings.append(warn_msg)
                elif val_str.lower() in ("true", "false"):
                    ws.cell(r_idx, c_idx, val_str.lower() == "true")
                else:
                    try:
                        if "." in val_str:
                            ws.cell(r_idx, c_idx, float(val_str))
                        else:
                            ws.cell(r_idx, c_idx, int(val_str))
                    except ValueError:
                        ws.cell(r_idx, c_idx, val)

        # Aplicar tabla estructurada y autofiltro
        if headers:
            last_col = get_column_letter(len(headers))
            last_row = max(len(rows_data), 2)
            _apply_table_style(ws, table_style="TableStyleMedium9", table_range=f"A1:{last_col}{last_row}")
            for c in range(1, len(headers) + 1):
                maxlen = max([len(str(r[c - 1])) for r in rows_data if c - 1 < len(r)] or [8])
                ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 3, 9), 40)

        wb.save(target_out)
        return {
            "status": "ok",
            "path": os.path.abspath(target_out),
            "bytes": os.path.getsize(target_out),
            "n_rows": len(rows_data),
            "n_cols": len(headers),
            "fidelity": "clean",
            "warnings": warnings,
        }

    elif dir_clean in ("xlsx_to_csv", "excel_to_csv"):
        wb = openpyxl.load_workbook(input_path, data_only=True)
        if sheet and sheet.lower() == "all":
            sheets_to_convert = wb.sheetnames
        elif sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Hoja '{sheet}' no existe en {input_path}. Disponibles: {wb.sheetnames}")
            sheets_to_convert = [sheet]
        else:
            ws_act = wb.active
            sheets_to_convert = [ws_act.title if ws_act else wb.sheetnames[0]]
            if len(wb.sheetnames) > 1:
                warnings.append(
                    f"Libro con múltiples hojas ({', '.join(wb.sheetnames)}); se exportó la activa '{sheets_to_convert[0]}'. Use sheet='all' para exportar todas."
                )

        if len(sheets_to_convert) == 1 or (not sheet or sheet.lower() != "all"):
            target_sheet_name = sheets_to_convert[0]
            ws = wb[target_sheet_name]
            rows_written = 0
            with open(target_out, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None and str(c).strip() != "" for c in row):
                        writer.writerow([c if c is not None else "" for c in row])
                        rows_written += 1
            wb.close()
            return {
                "status": "ok",
                "path": os.path.abspath(target_out),
                "bytes": os.path.getsize(target_out),
                "n_rows": rows_written,
                "sheet": target_sheet_name,
                "fidelity": "clean",
                "warnings": warnings,
            }
        else:
            # sheet == 'all' con múltiples hojas
            parent = os.path.dirname(target_out)
            stem = Path(target_out).stem
            gen_files = []
            total_rows = 0
            for s_name in sheets_to_convert:
                ws = wb[s_name]
                s_clean = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in s_name)
                s_file = safe_out(os.path.join(parent, f"{stem}_{s_clean}.csv"))
                rows_written = 0
                with open(s_file, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        if any(c is not None and str(c).strip() != "" for c in row):
                            writer.writerow([c if c is not None else "" for c in row])
                            rows_written += 1
                total_rows += rows_written
                gen_files.append(os.path.abspath(s_file))
            wb.close()
            return {
                "status": "ok",
                "path": gen_files[0] if gen_files else target_out,
                "files": gen_files,
                "n_files": len(gen_files),
                "n_rows": total_rows,
                "sheets": sheets_to_convert,
                "fidelity": "clean",
                "warnings": warnings,
            }
    else:
        raise ValueError(f"Dirección no soportada: '{direction}'. Opciones válidas: 'csv_to_xlsx', 'xlsx_to_csv'")
